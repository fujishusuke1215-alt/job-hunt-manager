"""Read only the public row values needed by the local personal-ranking importer."""
import json
import sys
from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit('xlsx path is required')
    workbook = load_workbook(sys.argv[1], data_only=True, read_only=True)
    if '総合ランキング' not in workbook.sheetnames:
        raise SystemExit('「総合ランキング」sheetがありません。')
    sheet = workbook['総合ランキング']
    rows = list(sheet.iter_rows(min_row=4, values_only=True))
    headers = [str(value or '') for value in rows[0]]
    records = [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]
    print(json.dumps(records, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()
